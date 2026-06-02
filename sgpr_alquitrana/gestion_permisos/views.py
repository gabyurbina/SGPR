"""Vistas principales de la aplicación SGPR.
Funciones para el registro de permisos, reposos, trabajadores y auditoría.
"""

from io import BytesIO
import datetime

from django.contrib import messages
from django.contrib.auth import logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.models import User
from django.contrib.auth import views as auth_views
from django.contrib.staticfiles import finders
from django.db.models import Q, Count
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
import json

from django.http import HttpResponse, JsonResponse
from django.conf import settings
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.utils import ImageReader
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as ReportLabImage

from .forms import (
    CedulaAuthenticationForm,
    CustomPasswordChangeForm,
    EliminarTrabajadorForm,
    EvaluacionSolicitudForm,
    RegistroTrabajadorForm,
    SolicitudForm,
    TrabajadorEditForm,
)
from .models import Auditoria, Solicitud, Trabajador
import logging
logger = logging.getLogger(__name__)
from .extra_views import vista_estadisticas


def get_encabezado_path():
    """Retorna la ruta absoluta del encabezado PNG si existe."""
    image_path = finders.find('images/encabezado.png') or finders.find('encabezado.png')
    return image_path


def format_datetime(dt):
    """Formato uniforme: DD-MM-YYYY hh:mm:ss a.m./p.m."""
    if not dt:
        return ''

    if isinstance(dt, datetime.date) and not isinstance(dt, datetime.datetime):
        return dt.strftime('%d-%m-%Y')

    formatted = dt.strftime('%d-%m-%Y %I:%M:%S %p')
    return formatted.replace('AM', 'a.m.').replace('PM', 'p.m.')


def agregar_numeracion_paginas(canvas, doc):
    """Dibuja encabezado (imagen, título y fecha) igual que la página 1 en todas las páginas, y número de página en el pie."""
    canvas.saveState()
    width, height = doc.pagesize

    # Encabezado: imagen en la esquina superior izquierda con tamaño máximo igual al de la primera página
    encabezado_path = get_encabezado_path()
    draw_h = 0
    if encabezado_path:
        try:
            image_reader = ImageReader(encabezado_path)
            orig_w, orig_h = image_reader.getSize()
            max_w, max_h = 420, 100
            scale = min(max_w / orig_w, max_h / orig_h, 1)
            draw_w = orig_w * scale
            draw_h = orig_h * scale
            x = doc.leftMargin
            y = height - draw_h - 20
            canvas.drawImage(encabezado_path, x, y, width=draw_w, height=draw_h, preserveAspectRatio=True, mask='auto')
        except Exception:
            draw_h = 0

    # Título centrado debajo del encabezado (mismo estilo que la primera página)
    title = getattr(doc, 'report_title', None)
    if title:
        canvas.setFont('Helvetica-Bold', 16)
        # Si hubo imagen, colocar título justo debajo; si no, usar un offset desde el tope
        if draw_h > 0:
            title_y = height - draw_h - 30
        else:
            title_y = height - 30
        canvas.drawCentredString(width / 2.0, title_y, title)

    # Fecha/datos a la derecha del título (misma línea que el título)
    gen = getattr(doc, 'generated_at', None)
    if gen:
        canvas.setFont('Helvetica', 9)
        # Ajustar ligeramente la posición vertical para alinearlo con el título
        if draw_h > 0:
            date_y = height - draw_h - 30
        else:
            date_y = height - 30
        canvas.drawRightString(width - doc.rightMargin, date_y, gen)

    # Número de página en el pie
    canvas.setFont('Helvetica', 8)
    page_number_text = f'Página {canvas.getPageNumber()}'
    canvas.drawRightString(width - doc.rightMargin, 15, page_number_text)
    canvas.restoreState()


def es_gestion_humana(user):
    """Determina si el usuario pertenece a Gestión Humana (staff o grupo)."""
    return user.is_staff or user.groups.filter(name='Gestion_Humana').exists()


@login_required
def dashboard(request):
    """Panel principal: lista de solicitudes según el rol del usuario."""
    query = request.GET.get('q', '').strip()
    filtro_estado = request.GET.get('estado', 'TODOS')
    tipo = request.GET.get('tipo', 'TODOS')
    active_tab = request.GET.get('tab', 'personal')

    is_worker = Trabajador.objects.filter(user=request.user).exists()

    # Solicitudes personales (si el usuario es trabajador)
    if request.user.is_authenticated and is_worker:
        personal_qs = Solicitud.objects.select_related('trabajador__user').filter(trabajador__user=request.user).order_by('-fecha_creacion')
    else:
        personal_qs = Solicitud.objects.none()

    # Solicitudes para administración (solo para gestores)
    admin_qs = None
    if es_gestion_humana(request.user):
        admin_qs = Solicitud.objects.select_related('trabajador__user').all().order_by('-fecha_creacion')

    # IDs de usuarios administradores que también son trabajadores.
    admin_worker_user_ids = []
    if es_gestion_humana(request.user):
        admin_worker_user_ids = list(
            User.objects.filter(
                pk__in=Trabajador.objects.values('user'),
            )
            .filter(
                Q(is_staff=True) | Q(groups__name='Gestion_Humana')
            )
            .distinct()
            .values_list('pk', flat=True)
        )

    # Aplicar filtros y búsqueda a cada queryset según corresponda
    def apply_filters(qs):
        if not qs:
            return qs
        if filtro_estado and filtro_estado != 'TODOS':
            qs = qs.filter(estado=filtro_estado)
        if tipo and tipo != 'TODOS':
            qs = qs.filter(tipo__iexact=tipo)
        if query:
            qs = qs.filter(
                Q(trabajador__cedula__icontains=query)
                | Q(trabajador__user__first_name__icontains=query)
                | Q(trabajador__user__last_name__icontains=query)
                | Q(trabajador__departamento__icontains=query)
                | Q(tipo__icontains=query)
                | Q(observaciones_admin__icontains=query)
            )
        return qs

    personal_qs = apply_filters(personal_qs)
    admin_qs = apply_filters(admin_qs) if admin_qs is not None else None

    # Paginación separada: `page_personal` y `page_admin`
    personal_page = None
    admin_page = None
    paginator_personal = None
    paginator_admin = None

    # Personal pagination
    page_personal = request.GET.get('page_personal', 1)
    paginator_personal = Paginator(personal_qs, 10)
    try:
        personal_page = paginator_personal.page(page_personal)
    except PageNotAnInteger:
        personal_page = paginator_personal.page(1)
    except EmptyPage:
        personal_page = paginator_personal.page(paginator_personal.num_pages)

    # Admin pagination (si aplica)
    if admin_qs is not None:
        page_admin = request.GET.get('page_admin', 1)
        paginator_admin = Paginator(admin_qs, 10)
        try:
            admin_page = paginator_admin.page(page_admin)
        except PageNotAnInteger:
            admin_page = paginator_admin.page(1)
        except EmptyPage:
            admin_page = paginator_admin.page(paginator_admin.num_pages)

    return render(
        request,
        'dashboard.html',
        {
            'personal_page': personal_page,
            'paginator_personal': paginator_personal,
            'admin_page': admin_page,
            'paginator_admin': paginator_admin,
            'es_admin': es_gestion_humana(request.user),
            'is_worker': is_worker,
            'is_super_admin': request.user.username == 'admin',
            'admin_worker_user_ids': admin_worker_user_ids,
            'active_tab': active_tab,
            'query': query,
            'filtro_estado': filtro_estado,
            'estado_options': [
                ('TODOS', 'Todos'),
                ('PENDIENTE', 'Pendientes'),
                ('APROBADO', 'Aprobadas'),
                ('RECHAZADO', 'Rechazadas'),
            ],
            'tipo_options': [
                ('TODOS', 'Todos'),
                ('PERMISO', 'Permiso'),
                ('REPOSO', 'Reposo'),
                ('OTRO', 'Otro'),
            ],
            'tipo': tipo,
        },
    )


def registro_trabajador(request):
    """Permite el registro de nuevos trabajadores desde la página pública o desde un administrador autenticado."""
    if request.user.is_authenticated and not es_gestion_humana(request.user):
        return redirect('dashboard')

    from_admin = request.user.is_authenticated and es_gestion_humana(request.user)
    if request.method == 'POST':
        form = RegistroTrabajadorForm(request.POST)
        if form.is_valid():
            cedula = form.cleaned_data['cedula']
            nombres = form.cleaned_data['nombres']
            apellidos = form.cleaned_data['apellidos']
            cargo = form.cleaned_data['cargo']
            departamento = form.cleaned_data['departamento']
            # Use default password and require change at first login
            password = settings.DEFAULT_PASSWORD
            email = form.cleaned_data.get('email')

            username = cedula
            user = User.objects.create_user(
                username=username,
                first_name=nombres,
                last_name=apellidos,
                password=password,
            )
            if email:
                user.email = email
                user.save()

            trabajador = Trabajador.objects.create(
                user=user,
                cedula=cedula,
                cargo=cargo,
                departamento=departamento,
                password_reset_required=True,
            )

            Auditoria.objects.create(
                usuario=request.user if request.user.is_authenticated else user,
                accion='REGISTRO',
                tabla_afectada='Trabajador',
                registro_id=trabajador.id,
                detalles=(
                    f"Registro de trabajador {trabajador.user.get_full_name()} (Cédula: {trabajador.cedula}) "
                    f"con usuario {user.username}."
                ),
            )

            show_password = settings.DEBUG or settings.SHOW_DEFAULT_PASSWORD_ON_REGISTER
            messages.success(
                request,
                'Trabajador registrado correctamente. Se ha creado la cuenta; el trabajador debe establecer su contraseña en el primer acceso.',
            )
            return render(
                request,
                'registro_exitoso.html',
                {
                    'username': username,
                    'from_admin': request.user.is_authenticated and es_gestion_humana(request.user),
                    'show_password': show_password,
                    'password': password if show_password else None,
                },
            )
    else:
        form = RegistroTrabajadorForm()

    return render(request, 'registro_trabajador.html', {'form': form, 'from_admin': from_admin})


@csrf_exempt
@require_POST
def api_registro_trabajador(request):
    """API para registrar un nuevo trabajador."""
    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
    except json.JSONDecodeError:
        payload = request.POST

    cedula = payload.get('cedula') or request.POST.get('cedula')
    nombres = payload.get('nombres') or request.POST.get('nombres')
    apellidos = payload.get('apellidos') or request.POST.get('apellidos')
    cargo = payload.get('cargo') or request.POST.get('cargo')
    departamento = payload.get('departamento') or request.POST.get('departamento')
    email = payload.get('email') or request.POST.get('email')

    if not all([cedula, nombres, apellidos, cargo, departamento]):
        return JsonResponse(
            {
                'success': False,
                'error': 'Faltan datos obligatorios. Se requieren cedula, nombres, apellidos, cargo y departamento.',
            },
            status=400,
        )

    if not cedula.isdigit():
        return JsonResponse(
            {
                'success': False,
                'error': 'La cédula solo debe contener números, sin letras ni caracteres especiales.',
            },
            status=400,
        )

    if Trabajador.objects.filter(cedula=cedula).exists() or User.objects.filter(username=cedula).exists():
        return JsonResponse(
            {'success': False, 'error': 'Ya existe un trabajador con esa cédula o usuario.'},
            status=400,
        )

    if email and User.objects.filter(email__iexact=email).exists():
        return JsonResponse(
            {'success': False, 'error': 'Ya existe un usuario registrado con este correo electrónico.'},
            status=400,
        )

    username = cedula
    password = settings.DEFAULT_PASSWORD
    user = User.objects.create_user(
        username=username,
        first_name=nombres,
        last_name=apellidos,
        password=password,
        email=email or '',
    )

    trabajador = Trabajador.objects.create(
        user=user,
        cedula=cedula,
        cargo=cargo,
        departamento=departamento,
        password_reset_required=True,
    )

    Auditoria.objects.create(
        usuario=request.user if request.user.is_authenticated else user,
        accion='REGISTRO',
        tabla_afectada='Trabajador',
        registro_id=trabajador.id,
        detalles=(
            f"Registro de trabajador {trabajador.user.get_full_name()} (Cédula: {trabajador.cedula}) "
            f"con usuario {user.username} mediante API."
        ),
    )

    return JsonResponse(
        {
            'success': True,
            'message': 'Trabajador registrado correctamente.',
            'username': username,
        }
    )


@login_required
@user_passes_test(es_gestion_humana)
def lista_trabajadores(request):
    """Listado de trabajadores para Gestión Humana."""
    query = request.GET.get('q', '').strip()
    filtro_estado = request.GET.get('estado', 'TODOS')
    trabajadores = Trabajador.objects.select_related('user').all().order_by('user__last_name')

    if filtro_estado == 'OBLIGATORIO':
        trabajadores = trabajadores.filter(password_reset_required=True)
    elif filtro_estado == 'COMPLETO':
        trabajadores = trabajadores.filter(password_reset_required=False)

    if query:
        trabajadores = trabajadores.filter(
            Q(cedula__icontains=query)
            | Q(user__first_name__icontains=query)
            | Q(user__last_name__icontains=query)
            | Q(user__email__icontains=query)
            | Q(cargo__icontains=query)
            | Q(departamento__icontains=query)
        )

    # Paginación: 10 trabajadores por página
    page = request.GET.get('page', 1)
    paginator = Paginator(trabajadores, 10)
    try:
        trabajadores_page = paginator.page(page)
    except PageNotAnInteger:
        trabajadores_page = paginator.page(1)
    except EmptyPage:
        trabajadores_page = paginator.page(paginator.num_pages)

    return render(
        request,
        'lista_trabajadores.html',
        {
            'trabajadores': trabajadores_page,
            'page_obj': trabajadores_page,
            'paginator': paginator,
            'query': query,
            'filtro_estado': filtro_estado,
            'estado_options': [
                ('TODOS', 'Todos'),
                ('OBLIGATORIO', 'Obligatorio'),
                ('COMPLETO', 'Completo'),
            ],
        },
    )


@login_required
@user_passes_test(es_gestion_humana)
def eliminar_trabajador(request, trabajador_id):
    """Eliminar un trabajador y registrar el motivo en auditoría."""
    trabajador = get_object_or_404(Trabajador, id=trabajador_id)
    if request.method == 'POST':
        form = EliminarTrabajadorForm(request.POST)
        if form.is_valid():
            motivo = form.cleaned_data['motivo'].strip()
            if not motivo:
                form.add_error('motivo', 'Debe indicar el motivo de la eliminación.')
            else:
                detalles = (
                    f"Trabajador eliminado: {trabajador.user.get_full_name()} ({trabajador.cedula}). "
                    f"Motivo: {motivo}"
                )
                Auditoria.objects.create(
                    usuario=request.user,
                    accion='ELIMINACIÓN',
                    tabla_afectada='Trabajador',
                    registro_id=trabajador.id,
                    detalles=detalles,
                )
                # Eliminar también el usuario asociado para no dejar huellas de acceso.
                trabajador.user.delete()
                messages.success(request, 'Trabajador eliminado correctamente y motivo registrado en auditoría.')
                return redirect('lista_trabajadores')
    else:
        form = EliminarTrabajadorForm()

    return render(
        request,
        'eliminar_trabajador.html',
        {
            'trabajador': trabajador,
            'form': form,
        },
    )


@login_required
@user_passes_test(es_gestion_humana)
def editar_trabajador(request, trabajador_id):
    """Permite a admin corregir datos del trabajador."""
    trabajador = get_object_or_404(Trabajador, id=trabajador_id)
    if request.method == 'POST':
        form = TrabajadorEditForm(request.POST, instance=trabajador)
        if form.is_valid():
            form.save()
            Auditoria.objects.create(
                usuario=request.user,
                accion='EDICIÓN',
                tabla_afectada='Trabajador',
                registro_id=trabajador.id,
                detalles=(
                    f"Datos actualizados para {trabajador.user.get_full_name()} (Cédula: {trabajador.cedula})."
                ),
            )
            messages.success(request, 'Datos del trabajador actualizados correctamente.')
            return redirect('lista_trabajadores')
    else:
        form = TrabajadorEditForm(instance=trabajador)

    return render(request, 'editar_trabajador.html', {'form': form, 'trabajador': trabajador})


@login_required
@user_passes_test(es_gestion_humana)
def restablecer_contrasena(request, trabajador_id):
    """Permite a admin reiniciar la contraseña de un trabajador."""
    trabajador = get_object_or_404(Trabajador, id=trabajador_id)
    if request.method == 'POST':
        trabajador.user.set_password(settings.DEFAULT_PASSWORD)
        trabajador.user.save()
        trabajador.password_reset_required = True
        trabajador.save()
        Auditoria.objects.create(
            usuario=request.user,
            accion='RESET_PASSWORD',
            tabla_afectada='User',
            registro_id=trabajador.user.id,
            detalles=(
                f"Contraseña restablecida para {trabajador.user.get_full_name()} (Cédula: {trabajador.cedula})."
            ),
        )
        messages.success(
            request,
            f"La contraseña de {trabajador.user.username} ha sido restablecida. El trabajador deberá cambiarla en su primer acceso.",
        )
        return redirect('lista_trabajadores')

    return render(request, 'restablecer_contrasena.html', {'trabajador': trabajador})


@login_required
def cambiar_contrasena(request):
    """Permite al usuario cambiar su contraseña y desactiva la obligación inicial."""
    if request.method == 'POST':
        form = CustomPasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            if hasattr(user, 'trabajador'):
                trabajador = user.trabajador
                trabajador.password_reset_required = False
                trabajador.save()
            Auditoria.objects.create(
                usuario=user,
                accion='CAMBIO_CONTRASENA',
                tabla_afectada='User',
                registro_id=user.id,
                detalles=(
                    f"Contraseña actualizada para {user.get_full_name() or user.username}."
                ),
            )
            from django.contrib import messages
            messages.success(request, 'Su contraseña fue cambiada con éxito.')
            return redirect('dashboard')
    else:
        form = CustomPasswordChangeForm(request.user)

    return render(request, 'cambiar_contrasena.html', {'form': form})


@login_required
def configuracion(request):
    """Página de configuración con opciones generales y administrativas."""
    # Si el usuario no es administrador, redirigimos directamente a cambiar_contrasena
    if not request.user.is_staff:
        return redirect('cambiar_contrasena')
    return render(request, 'configuracion.html', {'es_admin': request.user.is_staff})


@login_required
@user_passes_test(lambda u: u.is_staff)
def administrar_privilegios(request):
    """Permite a los administradores asignar o revocar privilegios de administrador."""
    query = request.GET.get('q', '').strip()
    is_admin = request.GET.get('is_admin', 'TODOS')
    accion = request.GET.get('accion', 'TODOS')
    trabajadores = Trabajador.objects.select_related('user').all().order_by('user__last_name')

    if is_admin == 'SI':
        trabajadores = trabajadores.filter(user__is_staff=True)
    elif is_admin == 'NO':
        trabajadores = trabajadores.filter(user__is_staff=False)

    if accion == 'grant':
        # Mostrar solo aquellos a los que se les puede otorgar (no son staff)
        trabajadores = trabajadores.filter(user__is_staff=False)
    elif accion == 'revoke':
        # Mostrar solo aquellos a los que se les puede revocar (son staff)
        trabajadores = trabajadores.filter(user__is_staff=True)

    if query:
        trabajadores = trabajadores.filter(
            Q(cedula__icontains=query)
            | Q(user__first_name__icontains=query)
            | Q(user__last_name__icontains=query)
            | Q(user__email__icontains=query)
            | Q(cargo__icontains=query)
            | Q(departamento__icontains=query)
        )

    # Paginación: 10 trabajadores por página
    page = request.GET.get('page', 1)
    paginator = Paginator(trabajadores, 10)
    try:
        trabajadores_page = paginator.page(page)
    except PageNotAnInteger:
        trabajadores_page = paginator.page(1)
    except EmptyPage:
        trabajadores_page = paginator.page(paginator.num_pages)

    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        action = request.POST.get('action')
        if user_id and action in ['grant', 'revoke']:
            try:
                trabajador = Trabajador.objects.select_related('user').get(user_id=int(user_id))
                target_user = trabajador.user
            except (Trabajador.DoesNotExist, ValueError):
                trabajador = None
                target_user = None

            if target_user is None:
                messages.error(request, 'Usuario no encontrado.')
            elif target_user == request.user and action == 'revoke':
                messages.error(request, 'No puedes quitarte privilegios a ti mismo desde esta página.')
            else:
                if action == 'grant':
                    target_user.is_staff = True
                    detalle = (
                        f'Privilegios de administrador otorgados a {target_user.get_full_name()} '
                        f'(Cédula: {trabajador.cedula}).'
                    )
                    messages.success(request, f'Se otorgaron privilegios a {target_user.get_full_name()}.')
                else:
                    target_user.is_staff = False
                    detalle = (
                        f'Privilegios de administrador revocados a {target_user.get_full_name()} '
                        f'(Cédula: {trabajador.cedula}).'
                    )
                    messages.success(request, f'Se revocaron privilegios a {target_user.get_full_name()}.')
                target_user.save()
                Auditoria.objects.create(
                    usuario=request.user,
                    accion='PRIVILEGIOS' if action == 'grant' else 'REVOCAR_PRIVILEGIOS',
                    tabla_afectada='User',
                    registro_id=target_user.id,
                    detalles=detalle,
                )
        else:
            messages.error(request, 'Solicitud inválida para la gestión de privilegios.')
        return redirect('privilegios')

    return render(
        request,
        'privilegios.html',
        {
            'trabajadores': trabajadores_page,
            'page_obj': trabajadores_page,
            'paginator': paginator,
            'query': query,
            'is_admin': is_admin,
            'accion': accion,
        },
    )


@login_required
def solicitar_permiso(request):
    """Permite al trabajador enviar una nueva solicitud con adjunto opcional."""
    if not Trabajador.objects.filter(user=request.user).exists() or request.user.username == 'admin':
        return redirect('dashboard')

    trabajador = get_object_or_404(Trabajador, user=request.user)
    if request.method == 'POST':
        form = SolicitudForm(request.POST, request.FILES)
        if form.is_valid():
            solicitud = form.save(commit=False)
            solicitud.trabajador = trabajador
            solicitud.save()
            Auditoria.objects.create(
                usuario=request.user,
                accion='CREACIÓN',
                tabla_afectada='Solicitud',
                registro_id=solicitud.id,
                detalles=(
                    f"El trabajador solicitó un {solicitud.tipo} desde {solicitud.fecha_inicio} "
                    f"hasta {solicitud.fecha_fin}."
                ),
            )
            return redirect('dashboard')
    else:
        form = SolicitudForm()

    return render(request, 'solicitar_permiso.html', {'form': form})


@login_required
@user_passes_test(es_gestion_humana)
def evaluar_solicitud(request, solicitud_id, accion):
    """Vista para aprobar o rechazar una solicitud con comentario obligatorio."""
    solicitud = get_object_or_404(Solicitud, id=solicitud_id)
    if accion not in ['aprobar', 'rechazar']:
        return redirect('dashboard')

    if request.method == 'POST':
        form = EvaluacionSolicitudForm(request.POST)
        if form.is_valid():
            observaciones = form.cleaned_data['observaciones_admin']
            solicitud.estado = 'APROBADO' if accion == 'aprobar' else 'RECHAZADO'
            solicitud.revisado_por = request.user
            solicitud.observaciones_admin = observaciones
            solicitud.save()
            Auditoria.objects.create(
                usuario=request.user,
                accion=accion.upper(),
                tabla_afectada='Solicitud',
                registro_id=solicitud.id,
                detalles=(
                    f"Solicitud ID {solicitud.id} de {solicitud.trabajador.user.get_full_name()} "
                    f"({solicitud.trabajador.cedula}) {solicitud.estado.lower()} por {request.user.get_full_name() or request.user.username}. "
                    f"Observaciones: {observaciones}"
                ),
            )
            return redirect('dashboard')
    else:
        form = EvaluacionSolicitudForm(initial={'observaciones_admin': solicitud.observaciones_admin})

    return render(
        request,
        'evaluar_solicitud.html',
        {
            'solicitud': solicitud,
            'accion': accion,
            'form': form,
        },
    )


@login_required
@user_passes_test(es_gestion_humana)
def reporte_auditoria(request):
    """Muestra los logs de auditoría al personal autorizado."""
    query = request.GET.get('q', '').strip()
    logs = Auditoria.objects.select_related('usuario').all()
    if query:
        q_filter = (
            Q(usuario__username__icontains=query)
            | Q(usuario__first_name__icontains=query)
            | Q(usuario__last_name__icontains=query)
            | Q(accion__icontains=query)
            | Q(tabla_afectada__icontains=query)
            | Q(detalles__icontains=query)
        )
        # If query is numeric, also filter by registro_id
        if query.isdigit():
            q_filter = q_filter | Q(registro_id=int(query))
        logs = logs.filter(q_filter)

    # Paginación: 10 logs por página
    page = request.GET.get('page', 1)
    paginator = Paginator(logs, 10)
    try:
        logs_page = paginator.page(page)
    except PageNotAnInteger:
        logs_page = paginator.page(1)
    except EmptyPage:
        logs_page = paginator.page(paginator.num_pages)

    return render(request, 'reporte_auditoria.html', {'logs': logs_page, 'page_obj': logs_page, 'paginator': paginator, 'query': query})


def add_excel_header_image(hoja, workbook):
    encabezado_path = get_encabezado_path()
    if encabezado_path:
        try:
            from openpyxl.drawing.image import Image as ExcelImage
            img = ExcelImage(encabezado_path)
            max_width, max_height = 520, 120
            try:
                from PIL import Image as PilImage
                with PilImage.open(encabezado_path) as pil_img:
                    orig_w, orig_h = pil_img.size
                    scale = min(max_width / orig_w, max_height / orig_h, 1)
                    img.width = int(orig_w * scale)
                    img.height = int(orig_h * scale)
            except ImportError:
                img.width = max_width
                img.height = max_height
            hoja.add_image(img, 'A1')
            hoja.row_dimensions[1].height = 90
        except Exception:
            pass


def add_pdf_header_image(documento, ancho, alto):
    encabezado_path = get_encabezado_path()
    if encabezado_path:
        try:
            image_reader = ImageReader(encabezado_path)
            orig_w, orig_h = image_reader.getSize()
            max_width, max_height = 420, 100
            scale = min(max_width / orig_w, max_height / orig_h, 1)
            draw_w = orig_w * scale
            draw_h = orig_h * scale
            x = 30
            y = alto - draw_h - 20
            documento.drawImage(encabezado_path, x, y, width=draw_w, height=draw_h, preserveAspectRatio=True, mask='auto')
        except Exception:
            pass


@login_required
@user_passes_test(es_gestion_humana)
def exportar_auditoria_excel(request):
    """Exporta el reporte de auditoría a Excel (respeta filtros de consulta)."""
    query = request.GET.get('q', '').strip()
    logs = Auditoria.objects.select_related('usuario').all()
    if query:
        q_filter = (
            Q(usuario__username__icontains=query)
            | Q(usuario__first_name__icontains=query)
            | Q(usuario__last_name__icontains=query)
            | Q(accion__icontains=query)
            | Q(tabla_afectada__icontains=query)
            | Q(detalles__icontains=query)
        )
        if query.isdigit():
            q_filter = q_filter | Q(registro_id=int(query))
        logs = logs.filter(q_filter)

    now = datetime.datetime.now()
    generated_at = now.strftime('%d-%m-%Y %I:%M:%S %p').replace('AM', 'a.m.').replace('PM', 'p.m.')
    timestamp = now.strftime('%Y%m%d_%I%M%S')

    workbook = Workbook()
    hoja = workbook.active
    hoja.title = 'Auditoría'
    add_excel_header_image(hoja, workbook)

    hoja.append(['Fecha de descarga:', '', '', '', '', generated_at])
    hoja.append([])
    headers = ['Fecha/Hora', 'Usuario', 'Acción', 'Tabla', 'Registro', 'Detalles']
    hoja.append(headers)

    for col_num, _ in enumerate(headers, start=1):
        cell = hoja.cell(row=3, column=col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for log in logs:
        hoja.append([
            format_datetime(log.fecha_hora),
            log.usuario_nombre_completo,
            log.accion,
            log.tabla_afectada,
            log.registro_id,
            log.detalles_legibles,
        ])

    for row in hoja.iter_rows(min_row=4, max_row=hoja.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    hoja.column_dimensions['A'].width = 20
    hoja.column_dimensions['B'].width = 30
    hoja.column_dimensions['C'].width = 20
    hoja.column_dimensions['D'].width = 20
    hoja.column_dimensions['E'].width = 18
    hoja.column_dimensions['F'].width = 60
    hoja.page_setup.fitToWidth = 1
    hoja.page_setup.fitToHeight = 0

    salida = BytesIO()
    workbook.save(salida)
    salida.seek(0)

    response = HttpResponse(
        salida.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename=auditoria_{timestamp}.xlsx'
    return response


@login_required
@user_passes_test(es_gestion_humana)
def exportar_auditoria_pdf(request):
    """Exporta el reporte de auditoría a PDF (respeta filtros de consulta)."""
    query = request.GET.get('q', '').strip()
    logs = Auditoria.objects.select_related('usuario').all()
    if query:
        q_filter = (
            Q(usuario__username__icontains=query)
            | Q(usuario__first_name__icontains=query)
            | Q(usuario__last_name__icontains=query)
            | Q(accion__icontains=query)
            | Q(tabla_afectada__icontains=query)
            | Q(detalles__icontains=query)
        )
        if query.isdigit():
            q_filter = q_filter | Q(registro_id=int(query))
        logs = logs.filter(q_filter)

    now = datetime.datetime.now()
    generated_at = now.strftime('%d-%m-%Y %I:%M:%S %p').replace('AM', 'a.m.').replace('PM', 'p.m.')
    timestamp = now.strftime('%Y%m%d_%I%M%S')

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=30, rightMargin=30, topMargin=90, bottomMargin=30)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        alignment=1,
        fontSize=16,
        leading=20,
        spaceAfter=10,
    )
    normal_style = ParagraphStyle(
        'Normal',
        parent=styles['Normal'],
        fontSize=10,
        leading=14,
    )

    story = []

    right_style = ParagraphStyle(
        'Right',
        parent=styles['Normal'],
        alignment=2,
        fontSize=9,
        leading=12,
        spaceAfter=12,
    )
    
    header_style = ParagraphStyle(
        'Header',
        parent=styles['Normal'],
        alignment=1,
        fontSize=10,
        leading=12,
        fontName='Helvetica-Bold',
        spaceAfter=4,
    )
    data = [[
        Paragraph('Fecha/Hora', header_style),
        Paragraph('Usuario', header_style),
        Paragraph('Acción', header_style),
        Paragraph('Tabla', header_style),
        Paragraph('Registro', header_style),
        Paragraph('Detalles', header_style),
    ]]

    cell_style = ParagraphStyle(
        'Cell',
        parent=styles['Normal'],
        alignment=0,
        fontSize=9,
        leading=12,
        spaceAfter=4,
    )

    for log in logs:
        detalles = Paragraph(str(log.detalles_legibles), cell_style)
        data.append([
            Paragraph(format_datetime(log.fecha_hora), cell_style),
            Paragraph(log.usuario_nombre_completo, cell_style),
            Paragraph(log.accion, cell_style),
            Paragraph(log.tabla_afectada, cell_style),
            Paragraph(str(log.registro_id), cell_style),
            detalles,
        ])

    table = Table(data, colWidths=[85, 115, 75, 75, 55, 150], repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f2f2f2')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('INNERGRID', (0, 0), (-1, -1), 0, colors.white),
        ('BOX', (0, 0), (-1, -1), 0, colors.white),
    ]))

    story.append(table)
    # Set metadata so el callback pueda dibujar título y fecha en cada página
    doc.report_title = 'Reporte de Auditoría'
    doc.generated_at = generated_at
    doc.build(story, onFirstPage=agregar_numeracion_paginas, onLaterPages=agregar_numeracion_paginas)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=auditoria_{timestamp}.pdf'
    return response


@login_required
@user_passes_test(es_gestion_humana)
def exportar_trabajadores_excel(request):
    """Exporta la lista de trabajadores a Excel (respeta filtros de consulta)."""
    query = request.GET.get('q', '').strip()
    filtro_estado = request.GET.get('estado', 'TODOS')
    trabajadores = Trabajador.objects.select_related('user').order_by('user__last_name')
    if filtro_estado == 'OBLIGATORIO':
        trabajadores = trabajadores.filter(password_reset_required=True)
    elif filtro_estado == 'COMPLETO':
        trabajadores = trabajadores.filter(password_reset_required=False)
    if query:
        trabajadores = trabajadores.filter(
            Q(cedula__icontains=query)
            | Q(user__first_name__icontains=query)
            | Q(user__last_name__icontains=query)
            | Q(user__email__icontains=query)
            | Q(cargo__icontains=query)
            | Q(departamento__icontains=query)
        )
    now = datetime.datetime.now()
    generated_at = now.strftime('%d-%m-%Y %I:%M:%S %p').replace('AM', 'a.m.').replace('PM', 'p.m.')
    workbook = Workbook()
    hoja = workbook.active
    hoja.title = 'Trabajadores'
    add_excel_header_image(hoja, workbook)

    hoja.append(['Fecha de descarga:', '', '', '', '', generated_at])
    hoja.append([])
    headers = ['Cédula', 'Nombre', 'Correo', 'Cargo', 'Ubicación Administrativa', 'Cambio clave']
    hoja.append(headers)

    for col_num, _ in enumerate(headers, start=1):
        cell = hoja.cell(row=3, column=col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for trabajador in trabajadores:
        hoja.append([
            trabajador.cedula,
            trabajador.user.get_full_name(),
            trabajador.user.email,
            trabajador.cargo,
            trabajador.departamento,
            'Obligatorio' if trabajador.password_reset_required else 'Completo',
        ])

    for row in hoja.iter_rows(min_row=4, max_row=hoja.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    hoja.column_dimensions['A'].width = 15
    hoja.column_dimensions['B'].width = 30
    hoja.column_dimensions['C'].width = 30
    hoja.column_dimensions['D'].width = 20
    hoja.column_dimensions['E'].width = 25
    hoja.column_dimensions['F'].width = 18
    hoja.page_setup.fitToWidth = 1
    hoja.page_setup.fitToHeight = 0

    salida = BytesIO()
    workbook.save(salida)
    salida.seek(0)
    response = HttpResponse(
        salida.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=trabajadores.xlsx'
    return response


@login_required
@user_passes_test(es_gestion_humana)
def exportar_trabajadores_pdf(request):
    """Exporta la lista de trabajadores a PDF (respeta filtros de consulta)."""
    query = request.GET.get('q', '').strip()
    filtro_estado = request.GET.get('estado', 'TODOS')
    trabajadores = Trabajador.objects.select_related('user').order_by('user__last_name')
    if filtro_estado == 'OBLIGATORIO':
        trabajadores = trabajadores.filter(password_reset_required=True)
    elif filtro_estado == 'COMPLETO':
        trabajadores = trabajadores.filter(password_reset_required=False)
    if query:
        trabajadores = trabajadores.filter(
            Q(cedula__icontains=query)
            | Q(user__first_name__icontains=query)
            | Q(user__last_name__icontains=query)
            | Q(user__email__icontains=query)
            | Q(cargo__icontains=query)
            | Q(departamento__icontains=query)
        )
    now = datetime.datetime.now()
    generated_at = now.strftime('%d-%m-%Y %I:%M:%S %p').replace('AM', 'a.m.').replace('PM', 'p.m.')
    timestamp = now.strftime('%Y%m%d_%I%M%S')

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=30, rightMargin=30, topMargin=90, bottomMargin=30)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        alignment=1,
        fontSize=16,
        leading=20,
        spaceAfter=10,
    )
    normal_style = ParagraphStyle(
        'Normal',
        parent=styles['Normal'],
        fontSize=10,
        leading=14,
    )
    header_style = ParagraphStyle(
        'Header',
        parent=styles['Normal'],
        alignment=1,
        fontSize=10,
        leading=12,
        fontName='Helvetica-Bold',
        spaceAfter=4,
    )
    cell_style = ParagraphStyle(
        'Cell',
        parent=styles['Normal'],
        alignment=0,
        fontSize=9,
        leading=12,
        spaceAfter=4,
    )

    story = []

    right_style = ParagraphStyle(
        'Right',
        parent=styles['Normal'],
        alignment=2,
        fontSize=9,
        leading=12,
        spaceAfter=12,
    )
    
    data = [[
        Paragraph('Cédula', header_style),
        Paragraph('Nombre', header_style),
        Paragraph('Correo', header_style),
        Paragraph('Cargo', header_style),
        Paragraph('Ubicación Administrativa', header_style),
        Paragraph('Cambio clave', header_style),
    ]]

    for trabajador in trabajadores:
        data.append([
            Paragraph(trabajador.cedula, cell_style),
            Paragraph(trabajador.user.get_full_name(), cell_style),
            Paragraph(trabajador.user.email or '', cell_style),
            Paragraph(trabajador.cargo or '', cell_style),
            Paragraph(trabajador.departamento or '', cell_style),
            Paragraph('Obligatorio' if trabajador.password_reset_required else 'Completo', cell_style),
        ])

    table = Table(data, colWidths=[70, 120, 110, 80, 95, 70], repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f2f2f2')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('INNERGRID', (0, 0), (-1, -1), 0, colors.white),
        ('BOX', (0, 0), (-1, -1), 0, colors.white),
    ]))

    story.append(table)
    # Set metadata so el callback pueda dibujar título y fecha en cada página
    doc.report_title = 'Lista de Trabajadores'
    doc.generated_at = generated_at
    doc.build(story, onFirstPage=agregar_numeracion_paginas, onLaterPages=agregar_numeracion_paginas)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=trabajadores_{timestamp}.pdf'
    return response


class LoginView(auth_views.LoginView):
    template_name = 'login.html'
    authentication_form = CedulaAuthenticationForm


def logout_view(request):
    """Cerrar sesión y redirigir al login con next=/"""
    logout(request)
    return redirect('/login/?next=/')


# --------------------- Estadísticas ---------------------
@login_required
@user_passes_test(es_gestion_humana)
def estadisticas(request):
    """Página de estadísticas con filtros y selector de tipo de gráfico."""
    # opciones y valores por defecto
    return render(request, 'estadisticas.html', {})


@login_required
@user_passes_test(es_gestion_humana)
def estadisticas_data(request):
    """Devuelve conteos agregados según filtros (JSON).
    Si se filtra por cédula (q), devuelve datasets por fecha y detalle de solicitudes."""
    q = request.GET.get('q', '').strip()
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    ubicacion = request.GET.get('ubicacion', '').strip()

    qs = Solicitud.objects.select_related('trabajador').all()
    if q:
        qs = qs.filter(trabajador__cedula__icontains=q)
    if ubicacion:
        qs = qs.filter(trabajador__departamento__icontains=ubicacion)
    if fecha_inicio:
        try:
            fi = datetime.datetime.strptime(fecha_inicio, '%Y-%m-%d')
            qs = qs.filter(fecha_creacion__date__gte=fi.date())
        except Exception:
            pass
    if fecha_fin:
        try:
            ff = datetime.datetime.strptime(fecha_fin, '%Y-%m-%d')
            qs = qs.filter(fecha_creacion__date__lte=ff.date())
        except Exception:
            pass

    # Si se aplicó cualquier filtro (cédula, fecha o ubicación) devolver datasets por fecha y detalle
    any_filter = bool(q or fecha_inicio or fecha_fin or ubicacion)
    if any_filter:
        # obtener fechas únicas ordenadas
        dates_qs = qs.dates('fecha_creacion', 'day')
        labels = [d.strftime('%d-%m-%Y') for d in dates_qs]
        statuses = [('APROBADO', '#1cc88a'), ('RECHAZADO', '#e74a3b'), ('PENDIENTE', '#f6c23e')]
        datasets = []
        for status, color in statuses:
            data = [qs.filter(estado=status, fecha_creacion__date=d).count() for d in dates_qs]
            datasets.append({'label': status, 'data': data, 'color': color})
        # detalle de solicitudes (fecha, estado, tipo, motivo)
        detail = []
        for s in qs.order_by('fecha_creacion'):
            detail.append({
                'fecha': s.fecha_creacion.strftime('%d-%m-%Y %I:%M:%S %p'),
                'estado': s.estado,
                'tipo': s.get_tipo_display() if hasattr(s, 'get_tipo_display') else s.tipo,
                'motivo': str(s.motivo) if s.motivo else ''
            })
        # también devolver desglose por ubicación
        loc_counts = list(qs.values('trabajador__departamento').annotate(c=Count('id')).order_by('-c'))
        loc_labels = [(d['trabajador__departamento'] or 'Sin Ubicación') for d in loc_counts[:10]]
        loc_values = [d['c'] for d in loc_counts[:10]]
        return JsonResponse({'labels': labels, 'datasets': datasets, 'detail': detail, 'locations': {'labels': loc_labels, 'values': loc_values}})

    # respuesta resumida (sin filtros)
    total = qs.count()
    aprobadas = qs.filter(estado='APROBADO').count()
    rechazadas = qs.filter(estado='RECHAZADO').count()
    pendientes = qs.filter(estado='PENDIENTE').count()

    labels = ['Total', 'Aprobadas', 'Rechazadas', 'Pendientes']
    values = [total, aprobadas, rechazadas, pendientes]

    # También devolver desglose por ubicación administrativa (top 10) para mostrar segundo gráfico en la UI
    loc_counts = list(qs.values('trabajador__departamento').annotate(c=Count('id')).order_by('-c'))
    loc_labels = [(d['trabajador__departamento'] or 'Sin Ubicación') for d in loc_counts[:10]]
    loc_values = [d['c'] for d in loc_counts[:10]]

    return JsonResponse({'labels': labels, 'values': values, 'locations': {'labels': loc_labels, 'values': loc_values}})


@login_required
@user_passes_test(es_gestion_humana)
def buscar_trabajadores(request):
    """Buscar trabajadores por nombre, apellido o cédula (devuelve JSON con id, nombre completo y cédula)."""
    q = request.GET.get('q', '').strip()
    results = []
    if q:
        try:
            workers_qs = Trabajador.objects.select_related('user').filter(
                Q(user__first_name__icontains=q) | Q(user__last_name__icontains=q) | Q(cedula__icontains=q)
            ).order_by('user__last_name')[:50]
            for t in workers_qs:
                results.append({'id': t.id, 'full_name': t.user.get_full_name(), 'cedula': t.cedula})
        except Exception:
            logger.exception('Error buscando trabajadores')
    return JsonResponse({'results': results})







@login_required
@user_passes_test(es_gestion_humana)
def exportar_estadisticas_pdf(request):
    """Exportar estadísticas a PDF (incrusta imagen del gráfico)."""
    q = request.GET.get('q', '').strip()
    trabajador_id = request.GET.get('trabajador_id')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    ubicacion = request.GET.get('ubicacion', '').strip()
    chart_type = request.GET.get('chart_type', 'pie')

    qs = Solicitud.objects.select_related('trabajador').all()
    # Prefer explicit trabajador_id (selection via search). Fallback to cedula query.
    if trabajador_id:
        try:
            qs = qs.filter(trabajador__id=int(trabajador_id))
        except Exception:
            pass
    else:
        if q:
            qs = qs.filter(trabajador__cedula__icontains=q)
    if ubicacion:
        qs = qs.filter(trabajador__departamento__icontains=ubicacion)
    if fecha_inicio:
        try:
            fi = datetime.datetime.strptime(fecha_inicio, '%Y-%m-%d')
            qs = qs.filter(fecha_creacion__date__gte=fi.date())
        except Exception:
            pass
    if fecha_fin:
        try:
            ff = datetime.datetime.strptime(fecha_fin, '%Y-%m-%d')
            qs = qs.filter(fecha_creacion__date__lte=ff.date())
        except Exception:
            pass

    total = qs.count()
    aprobadas = qs.filter(estado='APROBADO').count()
    rechazadas = qs.filter(estado='RECHAZADO').count()
    pendientes = qs.filter(estado='PENDIENTE').count()

    # Generar imágenes con matplotlib para PDF
    imgdata_requests = None
    imgdata_locations = None
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        any_filter = bool(q or fecha_inicio or fecha_fin or ubicacion)

        # Información para títulos
        fullname = ''
        if q and qs.exists():
            trabajador = qs.first().trabajador
            fullname = trabajador.user.get_full_name() if trabajador else ''
        range_text = ''
        if fecha_inicio or fecha_fin:
            range_text = f"{fecha_inicio or ''} - {fecha_fin or ''}"

        # Si hay filtros (por trabajador o fechas/ubicación) mostrar serie temporal si hay fechas, sino summary
        dates_qs = qs.dates('fecha_creacion', 'day') if qs.exists() else []
        dates_list = [d.strftime('%d-%m-%Y') for d in dates_qs]

        if any_filter and dates_qs:
            # stacked bar por fecha
            statuses = [('APROBADO', '#1cc88a'), ('RECHAZADO', '#e74a3b'), ('PENDIENTE', '#f6c23e')]
            fig, ax = plt.subplots(figsize=(6,1.6))
            bottoms = [0] * len(dates_list)
            for status, color in statuses:
                counts = [qs.filter(estado=status, fecha_creacion__date=d).count() for d in dates_qs]
                ax.bar(dates_list, counts, bottom=bottoms, label=status, color=color)
                bottoms = [a + b for a, b in zip(bottoms, counts)]
            title = 'Solicitudes por fecha'
            if fullname:
                title += f' - {fullname}'
            elif range_text:
                title += f' ({range_text})'
            ax.set_title(title, fontsize=10)
            ax.legend(loc='upper right', fontsize=8)
            ax.set_xticklabels(dates_list, rotation=45, ha='right', fontsize=7)
            ax.set_ylabel('Solicitudes')
            fig.tight_layout()
            imgbuf = BytesIO()
            fig.savefig(imgbuf, format='png', bbox_inches='tight', dpi=200)
            plt.close(fig)
            imgbuf.seek(0)
            imgdata_requests = imgbuf

            # ubicaciones
            loc_counts = list(qs.values('trabajador__departamento').annotate(c=Count('id')).order_by('-c'))
            loc_labels = [(d['trabajador__departamento'] or 'Sin Ubicación') for d in loc_counts[:10]]
            loc_values = [d['c'] for d in loc_counts[:10]]
            if loc_labels:
                fig2, ax2 = plt.subplots(figsize=(5,1.6))
                ax2.bar(loc_labels, loc_values, color=['#4e73df' if i % 2 == 0 else '#1cc88a' for i in range(len(loc_labels))])
                loc_title = 'Distribución por ubicación'
                if fullname:
                    loc_title += f' - {fullname}'
                ax2.set_title(loc_title, fontsize=10)
                ax2.set_xticklabels(loc_labels, rotation=45, ha='right', fontsize=7)
                ax2.set_ylabel('Solicitudes')
                fig2.tight_layout()
                imgbuf2 = BytesIO()
                fig2.savefig(imgbuf2, format='png', bbox_inches='tight', dpi=200)
                size_buf2b = len(imgbuf2.getvalue())
                logger.info(f'PDF: created imgbuf2 (locations) size={size_buf2b}')
                plt.close(fig2)
                imgbuf2.seek(0)
                imgdata_locations = imgbuf2
            else:
                imgdata_locations = None

        else:
            # Resumen global (no filtros relevantes): gráfico de estados y ubicaciones
            # Gráfico 1: resumen por estados (pie o bar según chart_type)
            if chart_type == 'pie':
                fig1, ax1 = plt.subplots(figsize=(3,1.4))
                mpl_colors = ['#1cc88a', '#e74a3b', '#f6c23e']
                ax1.pie([aprobadas, rechazadas, pendientes], labels=['Aprobadas', 'Rechazadas', 'Pendientes'], autopct='%1.1f%%', colors=mpl_colors)
                ax1.set_title('Resumen por estados', fontsize=10)
                ax1.legend(loc='lower center', bbox_to_anchor=(0.5, -0.3), ncol=3, fontsize=7)
            else:
                fig1, ax1 = plt.subplots(figsize=(5,1.6))
                status_labels = ['Aprobadas', 'Rechazadas', 'Pendientes']
                status_vals = [aprobadas, rechazadas, pendientes]
                ax1.bar(status_labels, status_vals, color=['#1cc88a', '#e74a3b', '#f6c23e'])
                ax1.set_title('Resumen por estados', fontsize=10)
                ax1.set_ylabel('Solicitudes')
            imgbuf1 = BytesIO()
            fig1.tight_layout()
            fig1.savefig(imgbuf1, format='png', bbox_inches='tight', dpi=200)
            size_buf1 = len(imgbuf1.getvalue())
            logger.info(f'PDF: created imgbuf1 size={size_buf1}')
            plt.close(fig1)
            imgbuf1.seek(0)
            imgdata_requests = imgbuf1

            # Gráfico 2: ubicaciones (top 10)
            loc_counts = list(qs.values('trabajador__departamento').annotate(c=Count('id')).order_by('-c'))
            loc_labels = [(d['trabajador__departamento'] or 'Sin Ubicación') for d in loc_counts[:10]]
            loc_values = [d['c'] for d in loc_counts[:10]]
            if loc_labels:
                fig2, ax2 = plt.subplots(figsize=(5,1.6))
                ax2.pie(loc_values, labels=loc_labels, autopct='%1.1f%%', colors=[('#4e73df' if i % 2 == 0 else '#1cc88a') for i in range(len(loc_labels))]) if chart_type == 'pie' else ax2.bar(loc_labels, loc_values, color=['#4e73df' if i % 2 == 0 else '#1cc88a' for i in range(len(loc_labels))])
                ax2.set_title('Distribución por ubicación', fontsize=10)
                fig2.tight_layout()
                imgbuf2 = BytesIO()
                fig2.savefig(imgbuf2, format='png', bbox_inches='tight', dpi=200)
                plt.close(fig2)
                imgbuf2.seek(0)
                imgdata_locations = imgbuf2
            else:
                imgdata_locations = None

    except Exception as e:
        logger.exception('PDF: error generando imágenes matplotlib')
        imgdata_requests = None
        imgdata_locations = None

    # Crear PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=30, rightMargin=30, topMargin=90, bottomMargin=30)
    styles = getSampleStyleSheet()
    story = []
    doc.report_title = 'Estadísticas de Solicitudes'
    doc.generated_at = datetime.datetime.now().strftime('%d-%m-%Y %I:%M:%S %p').replace('AM','a.m.').replace('PM','p.m.')

    # Añadir imágenes si existen: intentar colocarlas lado a lado si hay dos
    story.append(Spacer(1,12))
    pdf_temp_files = []
    try:
        import tempfile, os
        if imgdata_requests and imgdata_locations:
            # guardar ambos a archivos temporales y luego insertarlos por ruta
            imgdata_requests.seek(0)
            tmp_r = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
            try:
                from PIL import Image as PilImage
                pil_r = PilImage.open(imgdata_requests)
                if pil_r.mode in ('RGBA','LA'):
                    pil_r = pil_r.convert('RGB')
                pil_r.save(tmp_r.name, format='PNG')
                size_r = os.path.getsize(tmp_r.name)
                logger.info(f'PDF temp image created via PIL: {tmp_r.name} size={size_r}')
            except Exception:
                # fallback: write raw bytes
                logger.exception('PDF: PIL fallback writing tmp_r')
                with open(tmp_r.name, 'wb') as f:
                    f.write(imgdata_requests.getvalue())
            pdf_temp_files.append(tmp_r.name)
            imgdata_locations.seek(0)
            tmp_l = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
            try:
                from PIL import Image as PilImage
                pil_l = PilImage.open(imgdata_locations)
                if pil_l.mode in ('RGBA','LA'):
                    pil_l = pil_l.convert('RGB')
                pil_l.save(tmp_l.name, format='PNG')
                size_l = os.path.getsize(tmp_l.name)
                logger.info(f'PDF temp image created via PIL: {tmp_l.name} size={size_l}')
            except Exception:
                logger.exception('PDF: PIL fallback writing tmp_l')
                with open(tmp_l.name, 'wb') as f:
                    f.write(imgdata_locations.getvalue())
            pdf_temp_files.append(tmp_l.name)
            # scale images to available document width
            try:
                print('DEBUG: doc.width =', getattr(doc, 'width', 'UNKNOWN'))
                img_w = max((doc.width - 20) / 2.0, 100)
            except Exception as e:
                print('DEBUG: error computing img_w', e)
                img_w = 260
            print('DEBUG: writing images to', tmp_r.name, tmp_l.name)
            img_r = ReportLabImage(tmp_r.name, width=img_w)
            img_l = ReportLabImage(tmp_l.name, width=img_w)
            img_r.hAlign = 'CENTER'
            img_l.hAlign = 'CENTER'
            table_img = Table([[img_r, img_l]], colWidths=[img_w, img_w])
            table_img.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE')]))
            story.append(table_img)
            story.append(Spacer(1,12))
        else:
            if imgdata_requests:
                imgdata_requests.seek(0)
                tmp_r = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
                with open(tmp_r.name, 'wb') as f:
                    f.write(imgdata_requests.getvalue())
                try:
                    size_r = os.path.getsize(tmp_r.name)
                    if size_r > 0:
                        logger.info(f'PDF temp image created: {tmp_r.name} size={size_r}')
                    else:
                        logger.error(f'PDF temp image created but size=0: {tmp_r.name}')
                except Exception:
                    logger.exception('PDF: error obteniendo tamaño tmp_r')
                pdf_temp_files.append(tmp_r.name)
                try:
                    img_w = max(doc.width - 20, 200)
                except Exception:
                    img_w = 300
                img_r = ReportLabImage(tmp_r.name, width=img_w)
                img_r.hAlign = 'CENTER'
                story.append(img_r)
                story.append(Spacer(1,12))
            if imgdata_locations:
                imgdata_locations.seek(0)
                tmp_l = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
                try:
                    from PIL import Image as PilImage
                    pil_l = PilImage.open(imgdata_locations)
                    if pil_l.mode in ('RGBA','LA'):
                        pil_l = pil_l.convert('RGB')
                    pil_l.save(tmp_l.name, format='PNG')
                    size_l = os.path.getsize(tmp_l.name)
                    logger.info(f'PDF temp image created via PIL: {tmp_l.name} size={size_l}')
                except Exception:
                    logger.exception('PDF: PIL fallback writing tmp_l')
                    with open(tmp_l.name, 'wb') as f:
                        f.write(imgdata_locations.getvalue())
                pdf_temp_files.append(tmp_l.name)
                try:
                    img_w2 = max(doc.width - 20, 200)
                except Exception:
                    img_w2 = 300
                img_l = ReportLabImage(tmp_l.name, width=img_w2)
                img_l.hAlign = 'CENTER'
                story.append(img_l)
                story.append(Spacer(1,12))
    except Exception:
        logger.exception('PDF: fallo al insertar imágenes en el PDF')

    # Añadir tabla de resumen
    data = [['Métrica','Valor'], ['Total solicitudes', total], ['Aprobadas', aprobadas], ['Rechazadas', rechazadas], ['Pendientes', pendientes]]
    table = Table(data, colWidths=[200,100])
    story.append(table)

    # Si filtrado por trabajador, añadir encabezado con nombre y rango y detalle de solicitudes (fecha, tipo, estatus)
    if q or trabajador_id:
        # intentar obtener información del trabajador
        trabajador = None
        if qs.exists():
            trabajador = qs.first().trabajador
        fullname = trabajador.user.get_full_name() if trabajador else ''
        range_text = ''
        if fecha_inicio or fecha_fin:
            range_text = f'{fecha_inicio or ""} - {fecha_fin or ""}'
        # Agregar párrafo con información del trabajador
        try:
            from reportlab.platypus import Paragraph, TableStyle
            # Encabezado sin la etiqueta 'Rango' y con número de cédula
            cedula_val = trabajador.cedula if trabajador else ''
            header_style = getSampleStyleSheet()['Heading4']
            header_style.spaceAfter = 6
            detail_header = Paragraph(f'Reporte del trabajador: {fullname} — Cédula: {cedula_val}', header_style)
            story.append(Spacer(1,6))
            story.append(detail_header)
            story.append(Spacer(1,6))
        except Exception:
            pass

        story.append(Spacer(1,12))
        # Detalle de solicitudes: usar Paragraphs para que el texto se ajuste y no se sobreponga
        from reportlab.lib.styles import ParagraphStyle
        cell_style = ParagraphStyle('detail_cell', parent=styles['Normal'], fontSize=8, leading=10)
        header_style = ParagraphStyle('detail_header', parent=styles['Normal'], fontSize=9, leading=11, spaceAfter=4)
        detail_data = [[Paragraph('Fecha', header_style), Paragraph('Tipo', header_style), Paragraph('Estatus', header_style), Paragraph('Motivo', header_style)]]
        for s in qs.order_by('fecha_creacion'):
            tipo_label = s.get_tipo_display() if hasattr(s, 'get_tipo_display') else s.tipo
            fecha_txt = s.fecha_creacion.strftime('%d-%m-%Y %I:%M:%S %p')
            motivo_txt = str(s.motivo) if s.motivo else ''
            detail_data.append([
                Paragraph(fecha_txt, cell_style),
                Paragraph(tipo_label, cell_style),
                Paragraph(s.estado or '', cell_style),
                Paragraph(motivo_txt, cell_style),
            ])
        # calcular anchos de columna dinámicamente según ancho disponible
        try:
            avail_w = doc.width
        except Exception:
            from reportlab.lib.pagesizes import letter as _letter
            avail_w = _letter[0] - doc.leftMargin - doc.rightMargin
        col0 = 120
        col1 = 80
        col2 = 80
        col3 = max(avail_w - (col0 + col1 + col2), 120)
        detail_table = Table(detail_data, colWidths=[col0, col1, col2, col3], repeatRows=1)
        # Estilos: encabezado gris, grid, paddings, alineación y fondos alternos
        detail_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
            ('LEFTPADDING', (0,0), (-1,-1), 6),
            ('RIGHTPADDING', (0,0), (-1,-1), 6),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), (colors.white, colors.whitesmoke)),
        ]))
        story.append(Spacer(1,12))
        story.append(detail_table)

    doc.build(story, onFirstPage=agregar_numeracion_paginas, onLaterPages=agregar_numeracion_paginas)
    buffer.seek(0)
    # eliminar archivos temporales usados por PDF
    try:
        import os
        for _f in pdf_temp_files:
            try:
                os.remove(_f)
            except Exception:
                pass
    except Exception:
        pass
    timestamp = datetime.datetime.now().strftime('%Y%m%d_%I%M%S')
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=estadisticas_{timestamp}.pdf'
    return response

# ---------------------------------------------------------

